"""Build de-identified patient/image manifests and a private source map."""

from __future__ import annotations

import csv
import json
import secrets
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from src.common_paths import (  # noqa: E402
    PATIENT_MULTIMODAL_DERIVED_DIR,
    PATIENT_MULTIMODAL_REGISTRY_DIR,
    PATIENT_MULTIMODAL_REPORTS_DIR,
    RAW_LABEL_DIR,
)
from src.research_manifest import (  # noqa: E402
    extract_identity_token,
    pseudonymous_key,
    resolve_identity_tokens,
    validate_roi,
)
from src.research_schema import (  # noqa: E402
    DIAGNOSIS_TO_ID,
    EXCLUDED_DIAGNOSES,
    normalize_join_key,
)
from src.research_sources import IMAGE_EXTENSIONS  # noqa: E402


def write_csv(path: Path, rows: list[dict[str, object]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def load_excel_index(disease_dir: Path) -> tuple[str, dict[str, dict[str, object]]]:
    workbooks = [path for path in disease_dir.glob("*.xlsx") if not path.name.startswith("~$")]
    if not workbooks:
        return "", {}
    if len(workbooks) != 1:
        raise ValueError(f"Expected one workbook in {disease_dir}, found {len(workbooks)}")
    workbook = openpyxl.load_workbook(
        workbooks[0],
        read_only=True,
        data_only=True,
    )
    sheet = workbook[workbook.sheetnames[0]]
    index = {}
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        if not row or row[0] is None:
            continue
        key = normalize_join_key(row[0])
        if key in index:
            raise ValueError(f"Duplicate Excel join key in {workbooks[0]}: {key}")
        index[key] = {
            "row_number": row_number,
            "diagnosis_present": bool(len(row) > 3 and str(row[3] or "").strip()),
        }
    return workbooks[0].name, index


def identity_secret() -> bytes:
    private_dir = PATIENT_MULTIMODAL_REGISTRY_DIR / "private"
    private_dir.mkdir(parents=True, exist_ok=True)
    key_path = private_dir / "identity_hmac.key"
    if not key_path.exists():
        with key_path.open("xb") as handle:
            handle.write(secrets.token_bytes(32))
    return key_path.read_bytes()


def main() -> int:
    normalization = json.loads(
        (PATIENT_MULTIMODAL_REPORTS_DIR / "annotation_normalization.json").read_text(
            encoding="utf-8"
        )
    )
    normalized_root = (
        PATIENT_MULTIMODAL_DERIVED_DIR
        / normalization["annotation_version"]
        / "annotations"
    )
    secret = identity_secret()
    patient_rows = []
    image_rows = []
    private_identities = []
    private_sources = []
    excel_connections = []
    exclusions = Counter()
    identity_diagnoses = defaultdict(set)

    for disease_dir in sorted(path for path in RAW_LABEL_DIR.iterdir() if path.is_dir()):
        workbook_name, excel_index = load_excel_index(disease_dir)
        for patient_dir in sorted(path for path in disease_dir.iterdir() if path.is_dir()):
            jsons = {path.stem: path for path in patient_dir.glob("*.json")}
            images = {
                path.stem: path
                for path in patient_dir.iterdir()
                if path.is_file() and path.suffix.lower() in IMAGE_EXTENSIONS
            }
            paired_stems = sorted(jsons.keys() & images.keys())
            identities = set()
            loaded = {}
            for stem in paired_stems:
                annotation = json.loads(jsons[stem].read_text(encoding="utf-8"))
                loaded[stem] = annotation
                identities.add(extract_identity_token(annotation))

            identity_error = False
            try:
                canonical_identity = resolve_identity_tokens(
                    identities,
                    patient_folder=patient_dir.name,
                )
            except ValueError:
                identity_error = True
                canonical_identity = (
                    sorted(identities)[0]
                    if identities
                    else f"{disease_dir.name}:{patient_dir.name}"
                )
            person_key = pseudonymous_key(canonical_identity, secret, "KNEE_DEV")
            identity_diagnoses[person_key].add(disease_dir.name)

            join_key = normalize_join_key(patient_dir.name)
            excel = excel_index.get(join_key)
            included = disease_dir.name in DIAGNOSIS_TO_ID
            reasons = []
            if disease_dir.name in EXCLUDED_DIAGNOSES:
                included = False
                reasons.append("excluded_primary_diagnosis")
            if identity_error:
                included = False
                reasons.append("inconsistent_identity_within_patient")
            if not paired_stems:
                included = False
                reasons.append("no_paired_images")
            if disease_dir.name != "正常" and included and excel is None:
                included = False
                reasons.append("excel_unmatched")

            valid_roi = 0
            reviewed_roi = 0
            for stem in paired_stems:
                annotation = loaded[stem]
                info = annotation.get("info", {})
                width, height = info.get("width", 0), info.get("height", 0)
                roi_valid = validate_roi(
                    annotation.get("ultrasound_rect"),
                    width,
                    height,
                )
                reviewed = annotation.get("ultrasound_rect_reviewed") is True
                valid_roi += int(roi_valid)
                reviewed_roi += int(reviewed)
                image_key = pseudonymous_key(
                    f"{canonical_identity}|{stem}",
                    secret,
                    "KNEE_IMG",
                )
                rect = annotation.get("ultrasound_rect") or {}
                image_rows.append(
                    {
                        "image_key": image_key,
                        "person_key": person_key,
                        "diagnosis": disease_dir.name,
                        "diagnosis_id": DIAGNOSIS_TO_ID.get(disease_dir.name, ""),
                        "width": width,
                        "height": height,
                        "roi_x1": rect.get("x1", ""),
                        "roi_y1": rect.get("y1", ""),
                        "roi_x2": rect.get("x2", ""),
                        "roi_y2": rect.get("y2", ""),
                        "roi_valid": int(roi_valid),
                        "roi_reviewed": int(reviewed),
                        "annotation_object_count": len(annotation.get("objects", [])),
                        "include": int(included and roi_valid and reviewed),
                    }
                )
                normalized_json = (
                    normalized_root
                    / disease_dir.name
                    / patient_dir.name
                    / f"{stem}.json"
                )
                private_sources.append(
                    {
                        "image_key": image_key,
                        "raw_image_path": images[stem].relative_to(ROOT).as_posix(),
                        "raw_annotation_path": jsons[stem].relative_to(ROOT).as_posix(),
                        "normalized_annotation_path": normalized_json.relative_to(ROOT).as_posix(),
                    }
                )

            if valid_roi != len(paired_stems):
                included = False
                reasons.append("invalid_roi")
            if reviewed_roi != len(paired_stems):
                included = False
                reasons.append("roi_not_reviewed")
            for reason in reasons:
                exclusions[reason] += 1

            patient_rows.append(
                {
                    "person_key": person_key,
                    "cohort": "development_labeled",
                    "diagnosis": disease_dir.name,
                    "diagnosis_id": DIAGNOSIS_TO_ID.get(disease_dir.name, ""),
                    "include": int(included),
                    "exclusion_reason": ";".join(sorted(set(reasons))),
                    "clinical_available": int(excel is not None),
                    "paired_image_count": len(paired_stems),
                    "unannotated_image_count": len(images.keys() - jsons.keys()),
                    "orphan_annotation_count": len(jsons.keys() - images.keys()),
                    "valid_roi_count": valid_roi,
                    "reviewed_roi_count": reviewed_roi,
                }
            )
            private_identities.append(
                {
                    "person_key": person_key,
                    "diagnosis": disease_dir.name,
                    "patient_folder": patient_dir.name,
                    "identity_source_token": canonical_identity,
                }
            )
            excel_connections.append(
                {
                    "person_key": person_key,
                    "diagnosis": disease_dir.name,
                    "matched": int(excel is not None),
                    "workbook": workbook_name,
                    "row_number": "" if excel is None else excel["row_number"],
                    "diagnosis_present": (
                        "" if excel is None else int(excel["diagnosis_present"])
                    ),
                }
            )

    conflicts = {
        key: sorted(diagnoses)
        for key, diagnoses in identity_diagnoses.items()
        if len(diagnoses) > 1
    }
    if conflicts:
        for row in patient_rows:
            if row["person_key"] in conflicts:
                row["include"] = 0
                row["exclusion_reason"] = (
                    f"{row['exclusion_reason']};cross_diagnosis_identity_conflict"
                ).strip(";")
        for row in image_rows:
            if row["person_key"] in conflicts:
                row["include"] = 0
        exclusions["cross_diagnosis_identity_conflict"] = len(conflicts)

    patient_rows.sort(key=lambda row: (str(row["person_key"]), str(row["diagnosis"])))
    image_rows.sort(key=lambda row: str(row["image_key"]))
    private_sources.sort(key=lambda row: str(row["image_key"]))
    included_keys = {
        row["person_key"] for row in patient_rows if int(row["include"]) == 1
    }
    included_images = [
        row for row in image_rows if row["person_key"] in included_keys and row["include"] == 1
    ]

    registry = PATIENT_MULTIMODAL_REGISTRY_DIR
    private = registry / "private"
    write_csv(registry / "patients.csv", patient_rows, list(patient_rows[0]))
    write_csv(registry / "images.csv", image_rows, list(image_rows[0]))
    write_csv(registry / "excel_connections.csv", excel_connections, list(excel_connections[0]))
    write_csv(private / "person_identity_map.csv", private_identities, list(private_identities[0]))
    write_csv(private / "image_sources.csv", private_sources, list(private_sources[0]))

    diagnosis_counts = Counter(
        row["diagnosis"] for row in patient_rows if int(row["include"]) == 1
    )
    report = {
        "created_at": datetime.now().astimezone().isoformat(),
        "patient_rows": len(patient_rows),
        "included_patients": len(included_keys),
        "included_images": len(included_images),
        "diagnosis_patient_counts": dict(sorted(diagnosis_counts.items())),
        "all_included_images_have_valid_reviewed_roi": all(
            row["roi_valid"] == 1 and row["roi_reviewed"] == 1
            for row in included_images
        ),
        "excel_matches_included_abnormal": sum(
            int(row["matched"])
            for row in excel_connections
            if row["diagnosis"] != "正常"
            and row["person_key"] in included_keys
        ),
        "cross_diagnosis_identity_conflicts": conflicts,
        "exclusions": dict(sorted(exclusions.items())),
        "safe_manifests_contain_raw_paths": False,
        "private_identity_mapping": "private/person_identity_map.csv",
        "private_image_sources": "private/image_sources.csv",
    }
    (PATIENT_MULTIMODAL_REPORTS_DIR / "manifest_audit.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
