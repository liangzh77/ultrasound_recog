"""Leakage-resistant clinical feature preparation and five-fold baselines."""

from __future__ import annotations

import csv
import hashlib
import json
import math
import platform
import re
import subprocess
import time
from collections import Counter
from pathlib import Path
from typing import Any, Mapping, Sequence

import joblib
import numpy as np
import psutil
import yaml
from openpyxl import load_workbook
from sklearn.impute import SimpleImputer
from sklearn.linear_model import LogisticRegression
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler

from src.research_metrics import (
    bootstrap_macro_f1_ci,
    compute_patient_metrics,
    paired_bootstrap_macro_f1,
)
from src.research_schema import normalize_join_key
from src.research_tracking import LocalResearchTracker


CLINICAL_CLASSES = (
    "类风湿性关节炎",
    "痛风性关节炎",
    "脊柱关节炎",
    "骨性关节炎",
    "损伤",
)
ALLOWED_MODEL_FEATURES = (
    "sex",
    "age_years",
    "esr_mm_h",
    "crp_mg_l",
    "anti_ccp_u_ml",
    "rf_iu_ml",
    "hla_b27",
    "uric_acid",
)
LAB_FEATURES = (
    "esr_mm_h",
    "crp_mg_l",
    "anti_ccp_u_ml",
    "rf_iu_ml",
    "hla_b27",
    "uric_acid",
)
MISSING_FEATURES = tuple(f"{name}_missing" for name in LAB_FEATURES)
CENSOR_FEATURES = tuple(
    f"{name}_censor" for name in LAB_FEATURES if name != "hla_b27"
)
REFERENCE_COLUMNS = (
    "person_key",
    "reference_class",
    "reference_id",
    "outer_fold",
)
DERIVED_COLUMNS = (
    *REFERENCE_COLUMNS,
    *ALLOWED_MODEL_FEATURES,
    *MISSING_FEATURES,
    *CENSOR_FEATURES,
)
CLINICAL_PROBABILITY_COLUMNS = (
    "prob_ra",
    "prob_ga",
    "prob_spa",
    "prob_oa",
    "prob_injury",
)
FORBIDDEN_FEATURE_FRAGMENTS = (
    "诊断",
    "编号",
    "姓名",
    "病程",
    "日期",
    "路径",
    "文件名",
    "person_key",
    "reference",
)
SOURCE_COLUMN_INDEX = {
    "sex": 1,
    "age_years": 2,
    "esr_mm_h": 6,
    "crp_mg_l": 7,
    "anti_ccp_u_ml": 8,
    "rf_iu_ml": 9,
    "hla_b27": 10,
    "uric_acid": 11,
}
EXPECTED_HEADER_FRAGMENTS = (
    ("编号", "姓名"),
    ("性别",),
    ("年龄",),
    ("诊断",),
    ("病程",),
    ("超声检查日期",),
    ("血沉",),
    ("crp",),
    ("accp", "抗ccp", "抗-ccp"),
    ("rf",),
    ("hlab27", "hla-b27"),
    ("尿酸",),
)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_clinical_config(path: Path) -> dict[str, Any]:
    config = yaml.safe_load(path.read_text(encoding="utf-8"))
    if not isinstance(config, dict) or config.get("study_code") != "C0-C4":
        raise ValueError("Clinical config must define study_code C0-C4")
    if tuple(config.get("clinical_classes", ())) != CLINICAL_CLASSES:
        raise ValueError("Clinical class order differs from the frozen contract")
    contract = config.get("feature_contract", {})
    if tuple(contract.get("allowed_model_features", ())) != ALLOWED_MODEL_FEATURES:
        raise ValueError("Allowed clinical feature whitelist differs from ADR-002")
    if tuple(contract.get("laboratory_features", ())) != LAB_FEATURES:
        raise ValueError("Laboratory feature order differs from ADR-002")
    experiments = config.get("experiments", {})
    if tuple(experiments) != ("C0", "C1", "C2", "C3", "C4"):
        raise ValueError("Clinical experiments must be exactly C0 through C4")
    allowed_derived = set(ALLOWED_MODEL_FEATURES) | set(MISSING_FEATURES)
    for code, definition in experiments.items():
        features = tuple(definition.get("features", ()))
        if not set(features) <= allowed_derived:
            raise ValueError(f"{code} contains a non-whitelisted model feature")
        for feature in features:
            lowered = feature.casefold()
            if any(fragment.casefold() in lowered for fragment in FORBIDDEN_FEATURE_FRAGMENTS):
                raise ValueError(f"{code} contains forbidden feature: {feature}")
    if tuple(config.get("evaluation", {}).get("outer_folds", ())) != (0, 1, 2, 3, 4):
        raise ValueError("Clinical study requires the frozen five outer folds")
    if int(config.get("runtime", {}).get("max_cpu_threads", 0)) > 2:
        raise ValueError("Clinical study may use at most two CPU threads")
    if float(config.get("runtime", {}).get("hard_limit_hours_per_experiment", 99)) > 2:
        raise ValueError("Clinical experiment hard limit may not exceed two hours")
    return config


def parse_numeric_value(value: Any, missing_tokens: Sequence[str]) -> tuple[float | None, int, str | None]:
    """Return boundary value, censor code (-1/0/+1), and invalid token."""
    if value is None:
        return None, 0, None
    if isinstance(value, bool):
        return None, 0, str(value)
    if isinstance(value, (int, float)):
        numeric = float(value)
        return (numeric, 0, None) if math.isfinite(numeric) else (None, 0, str(value))
    text = str(value).strip()
    if text in set(missing_tokens):
        return None, 0, None
    text = text.replace("，", ",").replace("岁", "").strip()
    match = re.fullmatch(r"(?P<op><=|>=|<|>|≤|≥)?\s*(?P<number>[-+]?\d+(?:\.\d+)?)", text)
    if not match:
        return None, 0, str(value).strip()
    numeric = float(match.group("number"))
    operator = match.group("op")
    censor = -1 if operator in {"<", "<=", "≤"} else 1 if operator in {">", ">=", "≥"} else 0
    return numeric, censor, None


def parse_sex(value: Any) -> float:
    text = "" if value is None else str(value).strip().casefold()
    if text in {"男", "male", "m", "1"}:
        return 1.0
    if text in {"女", "female", "f", "0"}:
        return 0.0
    raise ValueError("Sex must be explicitly male or female")


def parse_hla_b27(value: Any, missing_tokens: Sequence[str]) -> tuple[float | None, str | None]:
    if value is None or str(value).strip() in set(missing_tokens):
        return None, None
    text = str(value).strip().casefold()
    if text in {"阳性", "positive", "+", "1"}:
        return 1.0, None
    if text in {"阴性", "negative", "-", "0"}:
        return 0.0, None
    return None, str(value).strip()


def validate_source_headers(headers: Sequence[str]) -> None:
    if len(headers) != len(EXPECTED_HEADER_FRAGMENTS):
        raise ValueError("Clinical source workbook must contain exactly 12 columns")
    for index, (header, allowed) in enumerate(zip(headers, EXPECTED_HEADER_FRAGMENTS)):
        normalized = re.sub(r"[\s_（）()／/]", "", str(header)).casefold()
        normalized_allowed = tuple(
            re.sub(r"[\s_（）()／/]", "", value).casefold() for value in allowed
        )
        if not any(fragment in normalized for fragment in normalized_allowed):
            raise ValueError(f"Clinical source column {index} does not match its frozen role")


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _write_csv(path: Path, fieldnames: Sequence[str], rows: Sequence[Mapping[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _workbook_inventory(raw_label_dir: Path, included_classes: Sequence[str]) -> dict[str, Path]:
    inventory: dict[str, Path] = {}
    for diagnosis in included_classes:
        files = sorted((raw_label_dir / diagnosis).glob("*.xlsx"))
        if len(files) != 1:
            raise ValueError(f"Expected exactly one workbook for included diagnosis: {diagnosis}")
        if files[0].name in inventory:
            raise ValueError("Workbook basenames must be unique")
        inventory[files[0].name] = files[0]
    return inventory


def _privacy_relative_path(path: Path) -> str:
    """Return a workspace-relative path without drive, user, or raw-source details."""
    parts = path.resolve().parts
    if "workspace" in parts:
        return Path(*parts[parts.index("workspace") :]).as_posix()
    return path.name


def prepare_clinical_features(
    config_path: Path,
    raw_label_dir: Path,
    registry_dir: Path,
    output_csv: Path,
    audit_json: Path,
) -> dict[str, Any]:
    config = load_clinical_config(config_path)
    missing_tokens = tuple(str(value) for value in config["feature_contract"]["accepted_missing_tokens"])
    patients = {
        row["person_key"]: row
        for row in _read_csv(registry_dir / "patients.csv")
        if int(row["include"]) == 1 and row["diagnosis"] in CLINICAL_CLASSES
    }
    folds = {row["person_key"]: int(row["outer_fold"]) for row in _read_csv(registry_dir / "folds_outer.csv")}
    private_map_path = registry_dir / "private" / "person_identity_map.csv"
    private_folders = {
        row["person_key"]: row["patient_folder"] for row in _read_csv(private_map_path)
    }
    connections = [
        row
        for row in _read_csv(registry_dir / "excel_connections.csv")
        if row["diagnosis"] in CLINICAL_CLASSES
    ]
    if len(patients) != 767 or len(connections) != 767:
        raise ValueError("Clinical v1 requires exactly 767 included abnormal patients")
    if {row["person_key"] for row in connections} != set(patients):
        raise ValueError("Excel connections do not cover the abnormal patient registry")
    if any(int(row["matched"]) != 1 for row in connections):
        raise ValueError("Every abnormal patient must have one matched Excel row")

    inventory = _workbook_inventory(raw_label_dir, CLINICAL_CLASSES)
    workbook_rows: dict[str, dict[int, dict[str, Any]]] = {}
    workbook_audit = []
    for basename, path in sorted(inventory.items()):
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        headers = tuple(
            "" if cell.value is None else str(cell.value).strip()
            for cell in next(sheet.iter_rows(min_row=1, max_row=1))
        )
        validate_source_headers(headers)
        rows = {}
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            safe_row = {
                feature: row[index] for feature, index in SOURCE_COLUMN_INDEX.items()
            }
            safe_row["join_key"] = row[0]
            rows[row_number] = safe_row
        workbook.close()
        workbook_rows[basename] = rows
        workbook_audit.append(
            {
                "workbook": basename,
                "sha256": sha256_file(path),
                "data_rows": len(rows),
                "columns": len(headers),
            }
        )

    invalid_counts: Counter[str] = Counter()
    censored_counts: Counter[str] = Counter()
    missing_counts: Counter[str] = Counter()
    derived_rows = []
    class_to_id = {name: index for index, name in enumerate(CLINICAL_CLASSES)}
    for connection in sorted(connections, key=lambda row: row["person_key"]):
        person_key = connection["person_key"]
        patient = patients[person_key]
        if connection["diagnosis"] != patient["diagnosis"]:
            raise ValueError("Connection and patient reference diagnosis differ")
        basename = connection["workbook"]
        if basename not in workbook_rows:
            raise ValueError("Connection references a workbook outside the included classes")
        row_number = int(connection["row_number"])
        if row_number not in workbook_rows[basename]:
            raise ValueError("Connection row is outside the source workbook")
        source = workbook_rows[basename][row_number]
        if person_key not in private_folders or normalize_join_key(
            source["join_key"]
        ) != normalize_join_key(private_folders[person_key]):
            raise ValueError("Excel row linkage no longer matches the private patient map")
        result: dict[str, Any] = {
            "person_key": person_key,
            "reference_class": patient["diagnosis"],
            "reference_id": class_to_id[patient["diagnosis"]],
            "outer_fold": folds[person_key],
            "sex": parse_sex(source["sex"]),
        }
        age, _, invalid_age = parse_numeric_value(
            source["age_years"], missing_tokens
        )
        if invalid_age is not None or age is None or not 0 < age <= 120:
            raise ValueError("Age must be numeric and within the frozen plausibility range")
        result["age_years"] = age

        for feature in LAB_FEATURES:
            raw_value = source[feature]
            if feature == "hla_b27":
                parsed, invalid = parse_hla_b27(raw_value, missing_tokens)
                censor = 0
            else:
                parsed, censor, invalid = parse_numeric_value(raw_value, missing_tokens)
                if parsed is not None and parsed < 0:
                    invalid = str(raw_value).strip()
                    parsed = None
                    censor = 0
            if invalid is not None:
                invalid_counts[feature] += 1
            missing = int(parsed is None)
            result[feature] = "" if parsed is None else parsed
            result[f"{feature}_missing"] = missing
            if feature != "hla_b27":
                result[f"{feature}_censor"] = censor
                censored_counts[feature] += int(censor != 0)
            missing_counts[feature] += missing
        derived_rows.append(result)

    _write_csv(output_csv, DERIVED_COLUMNS, derived_rows)
    output_sha = sha256_file(output_csv)
    audit = {
        "status": "PASS",
        "base_data_fingerprint": config["base_data_fingerprint"],
        "patients": len(derived_rows),
        "classes": list(CLINICAL_CLASSES),
        "class_counts": dict(Counter(row["reference_class"] for row in derived_rows)),
        "outer_fold_counts": dict(Counter(str(row["outer_fold"]) for row in derived_rows)),
        "source_workbooks": workbook_audit,
        "source_registry": {
            "patients_sha256": sha256_file(registry_dir / "patients.csv"),
            "folds_outer_sha256": sha256_file(registry_dir / "folds_outer.csv"),
            "excel_connections_sha256": sha256_file(registry_dir / "excel_connections.csv"),
            "private_identity_map_sha256": sha256_file(private_map_path),
        },
        "join_key_mismatches": 0,
        "model_feature_whitelist": list(ALLOWED_MODEL_FEATURES),
        "missingness_sensitivity_features": list(MISSING_FEATURES),
        "reference_only_columns": list(REFERENCE_COLUMNS),
        "excluded_source_roles": [
            "identity/linkage",
            "diagnosis/reference standard",
            "raw duration text",
            "date/batch",
            "path or filename",
        ],
        "missing_counts": dict(missing_counts),
        "censored_boundary_counts": dict(censored_counts),
        "nonnumeric_values_set_missing": dict(invalid_counts),
        "derived_table_path": _privacy_relative_path(output_csv),
        "derived_table_sha256": output_sha,
        "privacy": "No source identity, raw filename, raw path, diagnosis text, duration text, or date is copied into model features.",
    }
    audit_json.parent.mkdir(parents=True, exist_ok=True)
    audit_json.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    return audit


def load_clinical_table(path: Path) -> list[dict[str, str]]:
    rows = _read_csv(path)
    if not rows or tuple(rows[0]) != DERIVED_COLUMNS:
        raise ValueError("Clinical table columns differ from the frozen derived contract")
    keys = [row["person_key"] for row in rows]
    if len(keys) != len(set(keys)):
        raise ValueError("Clinical table person_key values must be unique")
    return rows


def _feature_matrix(
    rows: Sequence[Mapping[str, str]],
    features: Sequence[str],
    hla_missing_value: float,
) -> np.ndarray:
    matrix = []
    for row in rows:
        values = []
        for feature in features:
            value = row[feature]
            if value == "" and feature == "hla_b27":
                values.append(hla_missing_value)
            elif value == "":
                values.append(np.nan)
            else:
                values.append(float(value))
        matrix.append(values)
    return np.asarray(matrix, dtype=np.float64)


def build_logistic_pipeline(config: Mapping[str, Any], seed: int) -> Pipeline:
    model = config["model"]
    return Pipeline(
        steps=[
            ("imputer", SimpleImputer(strategy="median")),
            ("scaler", StandardScaler()),
            (
                "classifier",
                LogisticRegression(
                    solver=str(model["solver"]),
                    C=float(model["C"]),
                    class_weight=str(model["class_weight"]),
                    max_iter=int(model["max_iter"]),
                    random_state=seed,
                ),
            ),
        ]
    )


def _git_revision(project_root: Path) -> tuple[str, bool]:
    revision = subprocess.run(
        ["git", "rev-parse", "HEAD"], cwd=project_root, check=True, capture_output=True, text=True
    ).stdout.strip()
    dirty = bool(
        subprocess.run(
            ["git", "status", "--porcelain"],
            cwd=project_root,
            check=True,
            capture_output=True,
            text=True,
        ).stdout.strip()
    )
    return revision, dirty


def _write_clinical_oof(path: Path, rows: Sequence[Mapping[str, Any]]) -> None:
    fields = (
        "prediction_level",
        "person_key",
        "outer_fold",
        "reference_class",
        "reference_id",
        *CLINICAL_PROBABILITY_COLUMNS,
        "model_id",
    )
    _write_csv(path, fields, sorted(rows, key=lambda row: row["person_key"]))


def validate_clinical_oof(
    rows: Sequence[Mapping[str, str]], expected_keys: Sequence[str]
) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    if not rows:
        raise ValueError("Clinical OOF cannot be empty")
    keys = [row["person_key"] for row in rows]
    if len(keys) != len(set(keys)) or set(keys) != set(expected_keys):
        raise ValueError("Clinical OOF must cover every abnormal patient exactly once")
    if {row["prediction_level"] for row in rows} != {"patient_clinical"}:
        raise ValueError("Clinical OOF prediction level is invalid")
    targets = np.asarray([int(row["reference_id"]) for row in rows], dtype=np.int64)
    probabilities = np.asarray(
        [[float(row[column]) for column in CLINICAL_PROBABILITY_COLUMNS] for row in rows],
        dtype=np.float64,
    )
    folds = np.asarray([int(row["outer_fold"]) for row in rows], dtype=np.int64)
    if not np.isfinite(probabilities).all() or (probabilities < 0).any():
        raise ValueError("Clinical OOF probabilities must be finite and nonnegative")
    if not np.allclose(probabilities.sum(axis=1), 1.0, atol=1e-6):
        raise ValueError("Clinical OOF probabilities must sum to one")
    return targets, probabilities, folds


def evaluate_clinical_oof(
    path: Path,
    expected_keys: Sequence[str],
    bootstrap_samples: int,
    bootstrap_seed: int,
) -> dict[str, Any]:
    rows = _read_csv(path)
    targets, probabilities, folds = validate_clinical_oof(rows, expected_keys)
    interval = bootstrap_macro_f1_ci(
        targets, probabilities, n_bootstrap=bootstrap_samples, seed=bootstrap_seed
    )
    fold_metrics = {
        str(fold): compute_patient_metrics(
            targets[folds == fold], probabilities[folds == fold], CLINICAL_CLASSES
        )
        for fold in sorted(np.unique(folds))
    }
    return {
        "contract": {
            "prediction_level": "patient_clinical",
            "patients": len(rows),
            "unique_patients": len(set(row["person_key"] for row in rows)),
            "classes": list(CLINICAL_CLASSES),
            "outer_folds": sorted(int(value) for value in np.unique(folds)),
            "bootstrap_samples": bootstrap_samples,
            "bootstrap_seed": bootstrap_seed,
        },
        "prediction_sha256": sha256_file(path),
        "metrics": compute_patient_metrics(targets, probabilities, CLINICAL_CLASSES),
        "macro_f1_95_ci": list(interval),
        "fold_metrics": fold_metrics,
    }


def compare_clinical_oof(
    baseline_path: Path,
    candidate_path: Path,
    expected_keys: Sequence[str],
    bootstrap_samples: int,
    bootstrap_seed: int,
) -> dict[str, Any]:
    baseline_rows = _read_csv(baseline_path)
    candidate_rows = _read_csv(candidate_path)
    baseline_targets, baseline_probabilities, baseline_folds = validate_clinical_oof(
        baseline_rows, expected_keys
    )
    candidate_targets, candidate_probabilities, candidate_folds = validate_clinical_oof(
        candidate_rows, expected_keys
    )
    baseline_index = {row["person_key"]: index for index, row in enumerate(baseline_rows)}
    order = np.asarray([baseline_index[row["person_key"]] for row in candidate_rows])
    if not np.array_equal(baseline_targets[order], candidate_targets):
        raise ValueError("Paired clinical OOF reference targets differ")
    if not np.array_equal(baseline_folds[order], candidate_folds):
        raise ValueError("Paired clinical OOF outer folds differ")
    return {
        "comparison": "candidate_minus_baseline_macro_f1",
        "baseline_sha256": sha256_file(baseline_path),
        "candidate_sha256": sha256_file(candidate_path),
        "bootstrap_samples": bootstrap_samples,
        "bootstrap_seed": bootstrap_seed,
        **paired_bootstrap_macro_f1(
            candidate_targets,
            baseline_probabilities[order],
            candidate_probabilities,
            n_bootstrap=bootstrap_samples,
            seed=bootstrap_seed,
        ),
    }


def run_clinical_baselines(
    config_path: Path,
    clinical_table_path: Path,
    experiment_dir: Path,
    project_root: Path,
) -> dict[str, Any]:
    config = load_clinical_config(config_path)
    rows = load_clinical_table(clinical_table_path)
    if len(rows) != 767:
        raise ValueError("Clinical v1 requires exactly 767 abnormal patients")
    clinical_sha = sha256_file(clinical_table_path)
    config_sha = sha256_file(config_path)
    revision, dirty = _git_revision(project_root)
    if dirty:
        raise RuntimeError("Formal clinical baselines require a clean Git worktree")
    evaluation = config["evaluation"]
    tracker = LocalResearchTracker(experiment_dir / "tracking", "patient-primary-diagnosis")
    expected_keys = [row["person_key"] for row in rows]
    memory = psutil.virtual_memory()
    hardware = {
        "processor": platform.processor() or platform.machine(),
        "logical_cpu_count": psutil.cpu_count(logical=True),
        "physical_cpu_count": psutil.cpu_count(logical=False),
        "memory_total_gb": memory.total / (1024**3),
        "memory_available_start_gb": memory.available / (1024**3),
        "max_cpu_threads": int(config["runtime"]["max_cpu_threads"]),
        "gpu_used": False,
    }
    results: dict[str, Any] = {}
    for code, definition in config["experiments"].items():
        start = time.perf_counter()
        features = tuple(definition["features"])
        all_predictions: list[dict[str, Any]] = []
        fold_runtimes = []
        model_hashes = []
        fold_run_ids = []
        parent_metadata = {
            "study_code": "C0-C4",
            "experiment_code": code,
            "clinical_data_sha256": clinical_sha,
            "config_sha256": config_sha,
            "git_revision": revision,
            "feature_count": len(features),
            "estimator": definition["estimator"],
            "max_cpu_threads": int(config["runtime"]["max_cpu_threads"]),
        }
        with tracker.parent_run(f"{code}-clinical-fivefold-formal", parent_metadata) as parent:
            parent_run_id = parent.info.run_id
            for fold, seed in zip(evaluation["outer_folds"], evaluation["seeds"]):
                fold_start = time.perf_counter()
                train_rows = [row for row in rows if int(row["outer_fold"]) != int(fold)]
                test_rows = [row for row in rows if int(row["outer_fold"]) == int(fold)]
                train_targets = np.asarray([int(row["reference_id"]) for row in train_rows])
                test_targets = np.asarray([int(row["reference_id"]) for row in test_rows])
                with tracker.fold_run(
                    f"{code}-fold{fold}-seed{seed}-formal",
                    {"outer_fold": fold, "seed": seed, "feature_count": len(features)},
                ) as child:
                    fold_run_ids.append(child.info.run_id)
                    artifact_dir = experiment_dir / "artifacts" / "clinical" / code
                    artifact_dir.mkdir(parents=True, exist_ok=True)
                    if code == "C0":
                        counts = np.bincount(train_targets, minlength=len(CLINICAL_CLASSES)).astype(float)
                        probabilities = np.repeat((counts / counts.sum())[None, :], len(test_rows), axis=0)
                        model_path = artifact_dir / f"fold_{fold}_class_frequency.json"
                        model_path.write_text(
                            json.dumps(
                                {"classes": list(CLINICAL_CLASSES), "probabilities": (counts / counts.sum()).tolist()},
                                ensure_ascii=False,
                                indent=2,
                            ),
                            encoding="utf-8",
                        )
                    else:
                        train_matrix = _feature_matrix(
                            train_rows,
                            features,
                            float(config["feature_contract"]["hla_b27_missing_value"]),
                        )
                        test_matrix = _feature_matrix(
                            test_rows,
                            features,
                            float(config["feature_contract"]["hla_b27_missing_value"]),
                        )
                        pipeline = build_logistic_pipeline(config, int(seed))
                        pipeline.fit(train_matrix, train_targets)
                        probabilities = pipeline.predict_proba(test_matrix)
                        if tuple(pipeline.named_steps["classifier"].classes_) != tuple(range(len(CLINICAL_CLASSES))):
                            raise ValueError("Classifier probability class order is invalid")
                        model_path = artifact_dir / f"fold_{fold}_model.joblib"
                        joblib.dump(
                            {"pipeline": pipeline, "features": features, "classes": CLINICAL_CLASSES},
                            model_path,
                            compress=3,
                        )
                    model_sha = sha256_file(model_path)
                    model_hashes.append(
                        {
                            "fold": fold,
                            "path": model_path.relative_to(project_root).as_posix(),
                            "sha256": model_sha,
                        }
                    )
                    tracker.client.set_tag(child.info.run_id, "model_sha256", model_sha)
                    metrics = compute_patient_metrics(test_targets, probabilities, CLINICAL_CLASSES)
                    tracker.log_metrics(
                        {
                            "macro_f1": metrics["macro_f1"],
                            "accuracy": metrics["accuracy"],
                            "macro_auc": metrics["macro_auc"],
                            "ece": metrics["ece"],
                            "multiclass_brier": metrics["multiclass_brier"],
                        }
                    )
                fold_runtimes.append((time.perf_counter() - fold_start) / 3600)
                if (time.perf_counter() - start) / 3600 > float(
                    config["runtime"]["hard_limit_hours_per_experiment"]
                ):
                    raise TimeoutError(f"{code} exceeded its frozen two-hour hard limit")
                model_id = f"{code}-fold{fold}-seed{seed}-formal"
                for row, target, probability in zip(test_rows, test_targets, probabilities):
                    output = {
                        "prediction_level": "patient_clinical",
                        "person_key": row["person_key"],
                        "outer_fold": fold,
                        "reference_class": CLINICAL_CLASSES[int(target)],
                        "reference_id": int(target),
                        "model_id": model_id,
                    }
                    output.update(
                        {column: float(value) for column, value in zip(CLINICAL_PROBABILITY_COLUMNS, probability)}
                    )
                    all_predictions.append(output)

        oof_path = experiment_dir / "reports" / "clinical" / "oof" / f"{code}_oof.csv"
        _write_clinical_oof(oof_path, all_predictions)
        report = evaluate_clinical_oof(
            oof_path,
            expected_keys,
            int(evaluation["bootstrap_samples"]),
            int(evaluation["bootstrap_seed"]),
        )
        tracker.client.log_metric(
            parent_run_id, "oof_macro_f1", report["metrics"]["macro_f1"]
        )
        tracker.client.log_metric(
            parent_run_id, "oof_macro_auc", report["metrics"]["macro_auc"]
        )
        tracker.client.log_metric(parent_run_id, "oof_ece", report["metrics"]["ece"])
        tracker.client.set_tag(parent_run_id, "oof_sha256", report["prediction_sha256"])
        summary = {
            "experiment_code": code,
            "status": "COMPLETED",
            "research_purpose": definition["purpose"],
            "features": list(features),
            "clinical_data_sha256": clinical_sha,
            "config_path": config_path.relative_to(project_root).as_posix(),
            "config_sha256": config_sha,
            "git_revision": revision,
            "git_dirty": dirty,
            "folds": list(evaluation["outer_folds"]),
            "seeds": list(evaluation["seeds"]),
            "fold_runtime_hours": fold_runtimes,
            "fold_status": ["COMPLETED"] * len(fold_runtimes),
            "runtime_hours": (time.perf_counter() - start) / 3600,
            "hardware": hardware,
            "outer_test_used_for_fitting_preprocessing": False,
            "oof_path": oof_path.relative_to(project_root).as_posix(),
            "oof_sha256": report["prediction_sha256"],
            "metrics": report["metrics"],
            "macro_f1_95_ci": report["macro_f1_95_ci"],
            "fold_metrics": report["fold_metrics"],
            "model_artifacts": model_hashes,
            "mlflow_parent_run_id": parent_run_id,
            "mlflow_fold_run_ids": fold_run_ids,
        }
        summary_path = experiment_dir / "reports" / "clinical" / f"{code}_summary.json"
        summary_path.parent.mkdir(parents=True, exist_ok=True)
        summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
        tracker.client.set_tag(parent_run_id, "summary_sha256", sha256_file(summary_path))
        results[code] = {"summary_path": summary_path, "summary": summary}

    oof_dir = experiment_dir / "reports" / "clinical" / "oof"
    comparisons = {
        "C1_minus_C0": compare_clinical_oof(
            oof_dir / "C0_oof.csv",
            oof_dir / "C1_oof.csv",
            expected_keys,
            int(evaluation["bootstrap_samples"]),
            int(evaluation["bootstrap_seed"]),
        ),
        "C2_minus_C0": compare_clinical_oof(
            oof_dir / "C0_oof.csv",
            oof_dir / "C2_oof.csv",
            expected_keys,
            int(evaluation["bootstrap_samples"]),
            int(evaluation["bootstrap_seed"]),
        ),
        "C3_minus_C1": compare_clinical_oof(
            oof_dir / "C1_oof.csv",
            oof_dir / "C3_oof.csv",
            expected_keys,
            int(evaluation["bootstrap_samples"]),
            int(evaluation["bootstrap_seed"]),
        ),
        "C4_minus_C3": compare_clinical_oof(
            oof_dir / "C3_oof.csv",
            oof_dir / "C4_oof.csv",
            expected_keys,
            int(evaluation["bootstrap_samples"]),
            int(evaluation["bootstrap_seed"]),
        ),
    }
    combined = {
        "study_code": "C0-C4",
        "clinical_data_sha256": clinical_sha,
        "config_sha256": config_sha,
        "git_revision": revision,
        "experiments": {
            code: {
                "summary_path": item["summary_path"].relative_to(project_root).as_posix(),
                "macro_f1": item["summary"]["metrics"]["macro_f1"],
                "macro_f1_95_ci": item["summary"]["macro_f1_95_ci"],
                "oof_sha256": item["summary"]["oof_sha256"],
            }
            for code, item in results.items()
        },
        "paired_comparisons": comparisons,
    }
    combined_path = experiment_dir / "reports" / "clinical" / "C0_C4_evaluation.json"
    combined_path.write_text(json.dumps(combined, ensure_ascii=False, indent=2), encoding="utf-8")
    results["combined_evaluation"] = {
        "path": combined_path,
        "sha256": sha256_file(combined_path),
        "report": combined,
    }
    return results
